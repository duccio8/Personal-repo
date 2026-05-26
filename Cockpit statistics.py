"""Cockpit statistics report generator.

Fetches visit data from the Matomo `Live.getLastVisitsDetails` endpoint and
produces an Excel workbook with three sheets:

    1) Unique Users     - all unique user IDs since go-live, with the total
                          count shown prominently at the top.
    2) Active by Area   - most active users for the selected date range,
                          split by `dimension1` (business area), with chart.
    3) Top Tools        - most visited tools during the selected date range,
                          with chart.

Secrets:
    The Matomo API token is NEVER hardcoded. Create a `.env` file next to
    this script with:

        MATOMO_TOKEN=your_real_token_here

    The `.env` file is git-ignored (see `.gitignore`).
"""

from __future__ import annotations

import os
import re
import sys
from collections import Counter, defaultdict
from datetime import date, timedelta
from typing import Iterable

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

MATOMO_URL = "https://ecbdata.matomo.cloud"
SITE_ID = 22

# "Go-live" anchor used for Sheet 1 (unique users since go-live).
GO_LIVE_DATE = date(2026, 1, 1)

# Reporting window used for Sheet 2 and Sheet 3.
REPORT_START = GO_LIVE_DATE
REPORT_END = date(2026, 5, 30)

PAGE_SIZE = 1000
REQUEST_TIMEOUT = 60  # seconds
OUTPUT_FILE = "cockpit_statistics.xlsx"

# Number of top entries shown in the ranking sheets / charts.
TOP_USERS_PER_AREA = 10
TOP_TOOLS = 20

# Tool name = first path segment after `/tools/` in the action URL.
TOOL_URL_PATTERN = re.compile(r"/tools/([^/?#]+)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Secret management
# ---------------------------------------------------------------------------

def load_token() -> str:
    """Load the Matomo auth token from environment / `.env` file."""
    load_dotenv()
    token = os.getenv("MATOMO_TOKEN")
    if not token:
        sys.exit(
            "ERROR: MATOMO_TOKEN is not set.\n"
            "Create a `.env` file next to this script with:\n"
            "    MATOMO_TOKEN=<your token>\n"
        )
    return token


# ---------------------------------------------------------------------------
# Matomo client
# ---------------------------------------------------------------------------

def iter_week_mondays(start: date, end: date) -> Iterable[date]:
    """Yield each Monday between `start` and `end` (inclusive)."""
    current = start - timedelta(days=start.weekday())
    while current <= end:
        yield current
        current += timedelta(weeks=1)


def fetch_visits_for_week(token: str, week_monday: date) -> list[dict]:
    """Fetch all visits for the week starting on `week_monday` (paginated)."""
    visits: list[dict] = []
    offset = 0
    while True:
        params = {
            "module": "API",
            "method": "Live.getLastVisitsDetails",
            "idSite": SITE_ID,
            "period": "week",
            "date": week_monday.isoformat(),
            "format": "JSON",
            "token_auth": token,
            "filter_limit": PAGE_SIZE,
            "filter_offset": offset,
        }
        resp = requests.get(
            f"{MATOMO_URL}/index.php",
            params=params,
            timeout=REQUEST_TIMEOUT,
        )
        resp.raise_for_status()
        batch = resp.json()
        if not batch:
            break
        visits.extend(batch)
        if len(batch) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    return visits


def fetch_visits_in_range(
    token: str, start: date, end: date
) -> list[dict]:
    """Fetch all visits whose `serverDate` falls within [start, end]."""
    collected: list[dict] = []
    for monday in iter_week_mondays(start, end):
        print(f"Fetching week starting {monday} ...", end=" ", flush=True)
        weekly = fetch_visits_for_week(token, monday)

        kept = 0
        for v in weekly:
            try:
                sd = date.fromisoformat(v.get("serverDate", ""))
            except ValueError:
                continue
            if start <= sd <= end:
                collected.append(v)
                kept += 1
        print(f"{len(weekly)} fetched, {kept} in range")
    return collected


# ---------------------------------------------------------------------------
# Analytics
# ---------------------------------------------------------------------------

def user_key(visit: dict) -> str | None:
    """Return a stable user identifier, preferring `userId` over `visitorId`."""
    return visit.get("userId") or visit.get("visitorId")


def extract_tool_from_url(url: str | None) -> str | None:
    """Return the tool name (first segment after `/tools/`) or None."""
    if not url:
        return None
    match = TOOL_URL_PATTERN.search(url)
    return match.group(1).lower() if match else None


def compute_unique_users(visits: list[dict]) -> list[str]:
    """Sorted list of distinct user identifiers across the given visits."""
    users = {user_key(v) for v in visits if user_key(v)}
    return sorted(users)


def compute_active_users_by_area(
    visits: list[dict],
    top_n: int = TOP_USERS_PER_AREA,
) -> dict[str, list[tuple[str, int]]]:
    """Top-N users per business area, ranked by number of visits."""
    counters: dict[str, Counter] = defaultdict(Counter)
    for v in visits:
        uid = user_key(v)
        if not uid:
            continue
        area = v.get("dimension1") or "(unknown)"
        counters[area][uid] += 1  # one record == one visit
    return {
        area: counter.most_common(top_n)
        for area, counter in sorted(counters.items())
    }


def compute_top_tools(
    visits: list[dict],
    top_n: int = TOP_TOOLS,
) -> list[tuple[str, int]]:
    """Top-N tools by pageview count across all visits."""
    counter: Counter[str] = Counter()
    for visit in visits:
        for action in visit.get("actionDetails", []):
            if action.get("type") != "action":
                continue  # ignore content impressions / interactions / events
            tool = extract_tool_from_url(action.get("url"))
            if tool:
                counter[tool] += 1
    return counter.most_common(top_n)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
TITLE_FONT = Font(bold=True, size=16)
BIG_NUMBER_FONT = Font(bold=True, size=28, color="305496")


def _autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        length = max(
            (len(str(c.value)) for c in col if c.value is not None),
            default=10,
        )
        ws.column_dimensions[col[0].column_letter].width = min(length + 2, max_width)


def _write_header(ws, row: int, headers: list[str]) -> None:
    for i, value in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=value)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def build_sheet_unique_users(ws, user_ids: list[str]) -> None:
    ws["A1"] = "Total unique users since go-live"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"(go-live: {GO_LIVE_DATE.isoformat()})"
    ws["A2"].font = Font(italic=True, color="555555")

    ws["C1"] = len(user_ids)
    ws["C1"].font = BIG_NUMBER_FONT
    ws["C1"].alignment = Alignment(horizontal="center")

    _write_header(ws, 4, ["Unique User ID"])
    for i, uid in enumerate(user_ids, start=5):
        ws.cell(row=i, column=1, value=uid)

    ws.freeze_panes = "A5"
    _autosize(ws)


def build_sheet_active_users(
    ws, data: dict[str, list[tuple[str, int]]]
) -> None:
    ws["A1"] = "Most Active Users by Business Area"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Period: {REPORT_START.isoformat()} - {REPORT_END.isoformat()}"
    ws["A2"].font = Font(italic=True, color="555555")

    _write_header(ws, 4, ["Business Area", "User ID", "Visits"])

    row = 5
    for area, users in data.items():
        for uid, visits in users:
            ws.cell(row=row, column=1, value=area)
            ws.cell(row=row, column=2, value=uid)
            ws.cell(row=row, column=3, value=visits)
            row += 1

    last_row = row - 1
    ws.freeze_panes = "A5"
    _autosize(ws)

    if last_row < 5:
        return

    chart = BarChart()
    chart.type = "bar"
    chart.style = 11
    chart.title = "Top users by visits (grouped by business area)"
    chart.y_axis.title = "User (area)"
    chart.x_axis.title = "Visits"
    data_ref = Reference(ws, min_col=3, min_row=4, max_row=last_row, max_col=3)
    cat_ref = Reference(ws, min_col=1, min_row=5, max_row=last_row, max_col=2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.height = max(10, (last_row - 4) * 0.4)
    chart.width = 22
    chart.legend = None
    ws.add_chart(chart, "E4")


def build_sheet_top_tools(ws, tools: list[tuple[str, int]]) -> None:
    ws["A1"] = "Most Visited Tools"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Period: {REPORT_START.isoformat()} - {REPORT_END.isoformat()}"
    ws["A2"].font = Font(italic=True, color="555555")

    _write_header(ws, 4, ["Tool", "Pageviews"])
    for i, (tool, count) in enumerate(tools, start=5):
        ws.cell(row=i, column=1, value=tool)
        ws.cell(row=i, column=2, value=count)

    last_row = 4 + len(tools)
    ws.freeze_panes = "A5"
    _autosize(ws)

    if not tools:
        return

    chart = BarChart()
    chart.type = "bar"
    chart.style = 12
    chart.title = "Top tools by pageviews"
    chart.y_axis.title = "Tool"
    chart.x_axis.title = "Pageviews"
    data_ref = Reference(ws, min_col=2, min_row=4, max_row=last_row)
    cat_ref = Reference(ws, min_col=1, min_row=5, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.height = max(10, len(tools) * 0.45)
    chart.width = 22
    chart.legend = None
    ws.add_chart(chart, "D4")


def build_workbook(
    unique_users: list[str],
    active_by_area: dict[str, list[tuple[str, int]]],
    top_tools: list[tuple[str, int]],
    output_path: str,
) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Unique Users"
    build_sheet_unique_users(ws1, unique_users)

    ws2 = wb.create_sheet("Active Users by Area")
    build_sheet_active_users(ws2, active_by_area)

    ws3 = wb.create_sheet("Top Tools")
    build_sheet_top_tools(ws3, top_tools)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    token = load_token()

    # Sheet 1 spans go-live -> REPORT_END.
    # Sheets 2 and 3 span REPORT_START -> REPORT_END.
    # Fetch the widest window once and slice in-memory.
    fetch_start = min(GO_LIVE_DATE, REPORT_START)
    fetch_end = REPORT_END

    print(f"Fetching visits from {fetch_start} to {fetch_end} ...")
    all_visits = fetch_visits_in_range(token, fetch_start, fetch_end)
    print(f"\nTotal visits collected: {len(all_visits)}")

    visits_since_golive = [
        v for v in all_visits
        if date.fromisoformat(v["serverDate"]) >= GO_LIVE_DATE
    ]
    visits_in_report = [
        v for v in all_visits
        if REPORT_START <= date.fromisoformat(v["serverDate"]) <= REPORT_END
    ]

    unique_users = compute_unique_users(visits_since_golive)
    active_by_area = compute_active_users_by_area(visits_in_report)
    top_tools = compute_top_tools(visits_in_report)

    print(f"Unique users since go-live: {len(unique_users)}")
    print(f"Business areas in report:   {len(active_by_area)}")
    print(f"Distinct tools in report:   {len(top_tools)}")

    build_workbook(unique_users, active_by_area, top_tools, OUTPUT_FILE)
    print(f"\nReport written to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
